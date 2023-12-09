import bpy # type: ignore
import os 
import openpyxl # type: ignore
from bpy_extras.io_utils import ImportHelper # type: ignore

class ObjType(bpy.types.Operator):
    def is_mesh(scene, obj):
        return obj.type == "MESH"
    
    def is_armature(scene, obj):
        return obj.type == "ARMATURE"

class O_ImportSimpleExcel(bpy.types.Operator, ImportHelper):
    bl_idname = "excel.import_simple"
    bl_label = "导入Excel"
    filename_ext = ".xlsx"

    def execute(self, context):
        excel_file = self.filepath
        simple_main_column = context.scene.simple_main_column - 1 #excel从0开始数列
        simple_save_column = context.scene.simple_save_column - 1
        simple_toactive_column = context.scene.simple_toactive_column - 1
        simple_active_column = context.scene.simple_active_column - 1

        if not excel_file or not os.path.exists(excel_file):
            self.report({'ERROR'}, "请选择有效的Excel文件")
            return {'CANCELLED'}

        try:
            # 打开Excel文件
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # 读取Excel文件中的数据，并更新列表,min_row表示从第二行开始
            for row in sheet.iter_rows(min_row=2, values_only=True):
                value = str(row[simple_main_column])
                if (not value) or (value == "None"):
                    continue
                O_BoneSimpleMapping.bone_main.append(value)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                value = str(row[simple_save_column])
                if (not value) or (value == "None"):
                    continue
                O_BoneSimpleMapping.bone_save.append(value)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                key = str(row[simple_toactive_column])
                value = str(row[simple_active_column])
                if (not key) or (key == "None"):
                    continue
                O_BoneSimpleMapping.bone_mapping[key] = value

            
            self.report({'INFO'}, f"Excel文件已导入{excel_file}")
            return {'FINISHED'}
        except Exception as e:
            self.report({'ERROR'}, f"导入Excel文件时出现错误: {e}")
            return {'CANCELLED'}

class O_BoneSimpleMapping(bpy.types.Operator):
    bl_idname = "bone.simple_mapping"
    bl_label = "简化骨骼"
    bl_description = ""
    # 保存骨骼名称
    bone_main = []
    bone_save = []
    bone_mapping = {}
    
    def execute(self, context):
        if not self.bone_main:
            self.report({'ERROR'}, "似乎没有导入excel") 
            return {'FINISHED'}
        try:
            SourceArmature = bpy.data.objects.get(context.scene.simple_source_armature.name)
        except:
            self.report({'ERROR'}, "似乎没有选择对象") 
            return {'FINISHED'}
        bpy.ops.object.mode_set(mode='OBJECT')
        bpy.context.view_layer.objects.active = SourceArmature
        bpy.ops.object.mode_set(mode='POSE')
        bpy.ops.bone.pose_layers_merge() #调用合并骨骼层
        bpy.ops.pose.reveal() # 全部取消隐藏
        bpy.ops.pose.select_all(action='DESELECT')
        # 执行特定合并
        for simple_toactive_name, simple_active_name in self.bone_mapping.items():
            bpy.ops.pose.select_all(action='DESELECT')
            bpy.ops.object.select_pattern(pattern=simple_toactive_name)
            # 没有选择到骨骼则跳过
            if not context.selected_pose_bones:
                continue
            SourceArmature.data.bones.active = SourceArmature.data.bones[simple_active_name]
            try:
                bpy.ops.cats_manual.merge_weights_to_active()
            except Exception as e:
                self.report({'ERROR'}, f"调用cats时出现错误: {e}") 

        # 剩余合并至父级
        bpy.ops.pose.select_all(action='DESELECT')
        for bone in self.bone_main:
            bpy.ops.object.select_pattern(pattern=bone)
        for bone in self.bone_save:
            bpy.ops.object.select_pattern(pattern=bone)
        bpy.ops.pose.select_all(action='INVERT') # 反选
        try:
            bpy.ops.cats_manual.merge_weights()
        except Exception as e:
            self.report({'ERROR'}, f"调用cats时出现错误: {e}") 
        #删除剩余骨骼
        bpy.ops.object.mode_set(mode='EDIT')
        bpy.ops.armature.delete()
        bpy.ops.object.mode_set(mode='POSE')
        #全选并取消约束
        bpy.ops.pose.select_all(action='SELECT')
        bpy.ops.pose.constraints_clear()

        return {'FINISHED'}
        
class O_ImportPosExcel(bpy.types.Operator, ImportHelper):
    bl_idname = "excel.import_pos"
    bl_label = "导入Excel"
    filename_ext = ".xlsx"

    def execute(self, context):
        excel_file = self.filepath
        key_column = context.scene.key_column - 1
        value_column = context.scene.value_column - 1

        if not excel_file or not os.path.exists(excel_file):
            self.report({'ERROR'}, "请选择有效的Excel文件")
            return {'CANCELLED'}

        try:
            # 打开Excel文件
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # 清空现有的bone_mapping字典
            O_BonePosMapping.bone_mapping.clear()

            # 读取Excel文件中的数据，并更新bone_mapping字典
            for row in sheet.iter_rows(min_row=2, values_only=True):
                key = str(row[key_column])
                value = str(row[value_column])
                if (key == "None") or (value == "None"):
                    continue
                O_BonePosMapping.bone_mapping[key] = value

            self.report({'INFO'}, f"Excel文件已导入{excel_file}")
            return {'FINISHED'}
        except Exception as e:
            self.report({'ERROR'}, f"导入Excel文件时出现错误: {e}")
            return {'CANCELLED'}

class O_BonePosMapping(bpy.types.Operator):
    bl_idname = "bone.pos_mapping"
    bl_label = "复制位置"
    bl_description = "将源骨骼按excel对应, 添加复制位置约束并应用约束、骨架、姿态"
    # 获取要添加约束的源骨骼名称和目标骨骼名称的映射
    bone_mapping = {}
    
    def execute(self, context):
        if not self.bone_mapping:
            self.report({'ERROR'}, "似乎没有导入excel")
            return {'FINISHED'}
        try:
            SourceArmature = bpy.data.objects.get(context.scene.source_armature.name)
            TargetArmature = bpy.data.objects.get(context.scene.target_armature.name)
        except:
            self.report({'ERROR'}, "似乎没有选择对象") 
            return {'FINISHED'}
        # 遍历源骨骼名称映射
        for source_bone_name, target_bone_name in self.bone_mapping.items():
            # 获取源骨骼和目标骨骼
            source_bone = SourceArmature.pose.bones.get(source_bone_name)
            target_bone = TargetArmature.pose.bones.get(target_bone_name)
            
            # 检查是否找到了源骨骼和目标骨骼
            if source_bone and target_bone:
                # 为源骨骼添加“复制位置”约束
                constraint = source_bone.constraints.new('COPY_LOCATION')
                constraint.target = TargetArmature
                constraint.subtarget = target_bone_name

        # 应用约束
        bpy.context.view_layer.objects.active = SourceArmature
        bpy.ops.object.mode_set(mode='POSE')
        bpy.ops.pose.select_all(action='SELECT')
        for bone in SourceArmature.pose.bones:
            SourceArmature.data.bones.active = SourceArmature.data.bones[bone.name]
            # 遍历并应用骨骼上的约束
            for constraint in SourceArmature.pose.bones[bone.name].constraints:
                if constraint.type == "COPY_LOCATION":
                    bpy.ops.constraint.apply(constraint=constraint.name, owner='BONE')
        # 调用，应用骨架和姿态
        bpy.ops.bone.pose_apply()
        bpy.ops.object.mode_set(mode='OBJECT')     

        return {'FINISHED'}


class O_ImportRenameExcel(bpy.types.Operator, ImportHelper):
    bl_idname = "excel.import_rename"
    bl_label = "导入Excel"
    filename_ext = ".xlsx"

    def execute(self, context):
        excel_file = self.filepath
        rename_key_column = context.scene.rename_key_column - 1
        rename_value_column = context.scene.rename_value_column - 1
        rename_save_column = context.scene.rename_save_column - 1
        rename_target_save_column = context.scene.rename_target_save_column - 1

        if not excel_file or not os.path.exists(excel_file):
            self.report({'ERROR'}, "请选择有效的Excel文件")
            return {'CANCELLED'}

        try:
            # 打开Excel文件
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # 清空现有的bone_mapping字典
            O_BoneRenameMapping.bone_mapping.clear()

            # 读取Excel文件中的数据，并更新bone_mapping字典
            for row in sheet.iter_rows(min_row=2, values_only=True):
                key = str(row[rename_key_column])
                value = str(row[rename_value_column])
                if (key == "None") or (value == "None"):
                    continue
                O_BoneRenameMapping.bone_mapping[key] = value

            for row in sheet.iter_rows(min_row=2, values_only=True):
                value = str(row[rename_save_column])
                if (not value) or (value == "None"):
                    continue
                O_BoneRenameMapping.bone_save.append(value)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                value = str(row[rename_target_save_column])
                if (not value) or (value == "None"):
                    continue
                O_BoneRenameMapping.bone_target_save.append(value)

            self.report({'INFO'}, f"Excel文件已导入{excel_file}")
            return {'FINISHED'}
        except Exception as e:
            self.report({'ERROR'}, f"导入Excel文件时出现错误: {e}")
            return {'CANCELLED'}

class O_BoneRenameMapping(bpy.types.Operator):
    bl_idname = "bone.rename_mapping"
    bl_label = "重命名换绑"
    bl_description = "将源骨架骨骼按excel对应, 重命名至目标骨架并绑定"
    # 获取要添加约束的源骨骼名称和目标骨骼名称的映射
    bone_mapping = {}
    bone_save = []
    bone_save_parent = {}
    bone_target_save = []
    
    def execute(self, context):
        if not self.bone_mapping:
            self.report({'ERROR'}, "似乎没有导入excel")
            return {'FINISHED'}
        try:
            SourceMesh = bpy.data.objects.get(context.scene.rename_source_mesh.name)
            SourceArmature = bpy.data.objects.get(context.scene.rename_source_armature.name)
            TargetArmature = bpy.data.objects.get(context.scene.rename_target_armature.name)
        except:
            self.report({'ERROR'}, "似乎没有选择对象") 
            return {'FINISHED'}
        
        # 应用两个骨架姿态
        bpy.context.view_layer.objects.active = SourceArmature
        bpy.ops.object.mode_set(mode='POSE')
        bpy.ops.bone.pose_apply()
        bpy.ops.object.mode_set(mode='OBJECT')
        bpy.context.view_layer.objects.active = TargetArmature
        bpy.ops.object.mode_set(mode='POSE')
        bpy.ops.bone.pose_apply()

        # 姿态模式重命名
        bpy.ops.object.mode_set(mode='OBJECT')
        bpy.context.view_layer.objects.active = SourceArmature
        bpy.ops.object.mode_set(mode='POSE')
        for source_bone_name, target_bone_name in self.bone_mapping.items():
            #SourceArmature.data.bones.active = SourceArmature.data.bones[source_bone_name]
            #bpy.context.active_bone.name = target_bone_name
            try:
                SourceArmature.data.bones[source_bone_name].name = target_bone_name
            except:
                print(f"{source_bone_name}不存在")

        # 保留骨骼的父级记录
        for bone in self.bone_save:
            SourceArmature.data.bones.active = SourceArmature.data.bones[bone]
            bpy.ops.pose.select_hierarchy(direction='PARENT', extend=True)
            for bone_exist in self.bone_save:
                if context.active_pose_bone.name == bone_exist:
                    break
            else: # 遍历正常完成的最后执行
                key = bone
                value = context.active_pose_bone.name
                self.bone_save_parent[key] = value

        # 应用原骨架
        bpy.ops.object.mode_set(mode='OBJECT')
        bpy.context.view_layer.objects.active = SourceMesh
        for now_modifier in SourceMesh.modifiers:
            if now_modifier.type == 'ARMATURE':
                bpy.ops.object.modifier_apply(modifier=now_modifier.name) #应用
        # 取消父级保持变换结果
        bpy.ops.object.select_all(action='DESELECT')
        SourceArmature.select_set(True)
        SourceMesh.select_set(True)
        bpy.context.view_layer.objects.active = SourceArmature
        bpy.ops.object.parent_clear(type='CLEAR_KEEP_TRANSFORM')

               
        # 添加骨架修改器
        now_modifier = SourceMesh.modifiers.new(name=TargetArmature.name, type='ARMATURE')
        now_modifier.object = TargetArmature

        # 清理原骨架
        bpy.context.view_layer.objects.active = SourceArmature
        bpy.ops.object.mode_set(mode='EDIT')
        bpy.ops.armature.select_all(action='DESELECT')
        for bone in self.bone_save:
            bpy.ops.object.select_pattern(pattern=bone)
        bpy.ops.armature.select_all(action='INVERT') # 反选
        bpy.ops.armature.delete()
        # 添加骨架修改器
        now_modifier = SourceMesh.modifiers.new(name=TargetArmature.name, type='ARMATURE')
        now_modifier.object = TargetArmature

        # 清理原骨架
        bpy.context.view_layer.objects.active = SourceArmature
        bpy.ops.object.mode_set(mode='EDIT')
        bpy.ops.armature.select_all(action='DESELECT')
        for bone in self.bone_save:
            bpy.ops.object.select_pattern(pattern=bone)
        bpy.ops.armature.select_all(action='INVERT') # 反选
        bpy.ops.armature.delete()
       # 骨架合并
        bpy.ops.object.mode_set(mode='OBJECT')
        bpy.ops.object.select_all(action='DESELECT')
        SourceArmature.select_set(True) #bpy.ops.object.select_pattern(pattern=SourceArmature.name)
        TargetArmature.select_set(True) #bpy.ops.object.select_pattern(pattern=TargetArmature.name)
        bpy.context.view_layer.objects.active = TargetArmature
        bpy.ops.object.join()
        # 指定父级
        bpy.ops.object.select_all(action='DESELECT')
        SourceMesh.select_set(True)
        TargetArmature.select_set(True)
        bpy.context.view_layer.objects.active = TargetArmature
        bpy.ops.object.parent_set(type='OBJECT', keep_transform=True) 

        # 保留骨骼重新指定父级
        bpy.context.view_layer.objects.active = TargetArmature
        bpy.ops.object.mode_set(mode='EDIT')
        print(self.bone_save_parent)
        for bone_save_name, bone_parent_name in self.bone_save_parent.items():
            bpy.ops.armature.select_all(action='DESELECT') #取消选择
            bpy.ops.object.select_pattern(pattern=bone_save_name) #选择
            bpy.ops.object.select_pattern(pattern=bone_parent_name) #选择
            # 指定活动骨,取消相连项
            TargetArmature.data.edit_bones.active = TargetArmature.data.edit_bones[bone_save_name]
            context.active_bone.use_connect = False
            # 指定活动骨,创建父级
            TargetArmature.data.edit_bones.active = TargetArmature.data.edit_bones[bone_parent_name]
            bpy.ops.armature.parent_set(type='OFFSET')

        # 主骨骼层
        bpy.ops.object.mode_set(mode='POSE')
        bpy.ops.pose.select_all(action='DESELECT') #取消选择
        for source_bone_name, target_bone_name in self.bone_mapping.items():
            bpy.ops.object.select_pattern(pattern=target_bone_name)
        bpy.ops.pose.group_add()
        TargetArmature.pose.bone_groups.active.name = "主骨骼"
        TargetArmature.pose.bone_groups.active.color_set = 'THEME02'
        bpy.ops.pose.group_assign(type=1)
        # 源保留骨骼层
        bpy.ops.pose.select_all(action='DESELECT') #取消选择
        for bone in self.bone_save:
            bpy.ops.object.select_pattern(pattern=bone)
        bpy.ops.pose.group_add()
        TargetArmature.pose.bone_groups.active.name = "源保留骨骼"
        TargetArmature.pose.bone_groups.active.color_set = 'THEME09'
        bpy.ops.pose.group_assign(type=2)
        # 目标保留骨骼层
        bpy.ops.pose.select_all(action='DESELECT') #取消选择
        for bone in self.bone_target_save:
            bpy.ops.object.select_pattern(pattern=bone)
        bpy.ops.pose.group_add()
        TargetArmature.pose.bone_groups.active.name = "目标保留骨骼"
        TargetArmature.pose.bone_groups.active.color_set = 'THEME07'
        bpy.ops.pose.group_assign(type=3)
        # 其余骨骼层
        TargetArmature.pose.bone_groups.active_index = 0
        bpy.ops.pose.group_select()
        TargetArmature.pose.bone_groups.active_index = 1
        bpy.ops.pose.group_select()
        TargetArmature.pose.bone_groups.active_index = 2
        bpy.ops.pose.group_select()
        bpy.ops.pose.select_all(action='INVERT') # 反选
        bpy.ops.pose.group_add()
        TargetArmature.pose.bone_groups.active.name = "其余骨骼"
        bpy.ops.pose.group_assign(type=4)
        
        


        return {'FINISHED'}



class P_BoneMapping(bpy.types.Panel):
    bl_label = "MOD骨架替换"
    bl_idname = "PT_BoneMapping"
    bl_space_type = 'VIEW_3D'
    bl_region_type = 'UI'
    bl_category = 'Bone+'
    
    def draw(self, context):
        layout = self.layout
        
        box = layout.box()
        # 选择骨架
        box.prop(context.scene, "simple_source_armature", text="", icon="ARMATURE_DATA")
        # 导入Excel文件按钮和文本框
        col = box.column(align=True)
        row = col.row(align=True)
        row.prop(context.scene, "simple_main_column", text="主骨骼列")
        row.prop(context.scene, "simple_save_column", text="保留骨骼列")
        row = col.row(align=True)
        row.prop(context.scene, "simple_toactive_column", text="被合并骨骼列")
        row.prop(context.scene, "simple_active_column", text="合并骨骼列")
        row = col.row(align=True)
        row.operator(O_ImportSimpleExcel.bl_idname, icon="IMPORT")
        # 添加按钮
        row.operator(O_BoneSimpleMapping.bl_idname, icon="PLAY")


        box = layout.box()
        # 选择源骨架
        col = box.column(align=True)
        col.label(text="选择源骨架:")
        col.prop(context.scene, "source_armature", text="", icon="ARMATURE_DATA")
        # 选择目标骨架
        col = box.column(align=True)
        col.label(text="选择目标骨架:")
        col.prop(context.scene, "target_armature", text="", icon="ARMATURE_DATA")
        # 导入Excel文件按钮和文本框
        col = box.column(align=True)
        row = col.row(align=True)
        row.prop(context.scene, "key_column", text="源骨架列")
        row.prop(context.scene, "value_column", text="目标骨架列")
        row = col.row(align=True)
        row.operator(O_ImportPosExcel.bl_idname, icon="IMPORT")
        # 添加按钮
        row.operator(O_BonePosMapping.bl_idname, icon="PLAY")
        

        box = layout.box()
        # 选择源物体、源骨架
        col = box.column(align=True)
        col.label(text="选择源物体、源骨架:")
        col.prop(context.scene, "rename_source_mesh", text="", icon="MESH_DATA")
        col.prop(context.scene, "rename_source_armature", text="", icon="ARMATURE_DATA")
        # 选择目标骨架
        col = box.column(align=True)
        col.label(text="选择目标骨架:")
        col.prop(context.scene, "rename_target_armature", text="", icon="ARMATURE_DATA")
        # 导入Excel文件按钮和文本框
        col = box.column(align=True)
        row = col.row(align=True)
        row.prop(context.scene, "rename_key_column", text="源骨架列")
        row.prop(context.scene, "rename_value_column", text="目标骨架列")
        row = col.row(align=True)
        row.prop(context.scene, "rename_save_column", text="源保留骨骼列")
        row.prop(context.scene, "rename_target_save_column", text="目标保留骨骼列")
        row = col.row(align=True)
        row.operator(O_ImportRenameExcel.bl_idname, icon="IMPORT")
        # 添加按钮
        row.operator(O_BoneRenameMapping.bl_idname, icon="PLAY")



def register():
    bpy.utils.register_class(O_ImportSimpleExcel)
    bpy.utils.register_class(O_BoneSimpleMapping)
    bpy.utils.register_class(O_ImportPosExcel)
    bpy.utils.register_class(O_BonePosMapping)
    bpy.utils.register_class(O_ImportRenameExcel)
    bpy.utils.register_class(O_BoneRenameMapping)
    bpy.utils.register_class(P_BoneMapping)
    ########################## Divider ##########################
    bpy.types.Scene.simple_source_armature = bpy.props.PointerProperty(type=bpy.types.Object, poll=ObjType.is_armature)
    bpy.types.Scene.source_armature = bpy.props.PointerProperty(type=bpy.types.Object, poll=ObjType.is_armature)
    bpy.types.Scene.target_armature = bpy.props.PointerProperty(type=bpy.types.Object, poll=ObjType.is_armature)
    bpy.types.Scene.rename_source_mesh = bpy.props.PointerProperty(type=bpy.types.Object, poll=ObjType.is_mesh)
    bpy.types.Scene.rename_source_armature = bpy.props.PointerProperty(type=bpy.types.Object, poll=ObjType.is_armature)
    bpy.types.Scene.rename_target_armature = bpy.props.PointerProperty(type=bpy.types.Object, poll=ObjType.is_armature)
    ########################## Divider ##########################
    bpy.types.Scene.simple_main_column = bpy.props.IntProperty(
        name="主骨骼列",
        default=2,
        min=1,
    )
    bpy.types.Scene.simple_save_column = bpy.props.IntProperty(
        name="保留骨骼列",
        default=3,
        min=1,
    )

    bpy.types.Scene.simple_toactive_column = bpy.props.IntProperty(
        name="值列",
        default=4,
        min=1,
    )
    bpy.types.Scene.simple_active_column = bpy.props.IntProperty(
        name="键列",
        default=5,
        min=1,
    )
    ########################## Divider ##########################
    bpy.types.Scene.key_column = bpy.props.IntProperty(
        name="键列",
        default=1,
        min=1,
    )
    bpy.types.Scene.value_column = bpy.props.IntProperty(
        name="值列",
        default=2,
        min=1,
    )
    ########################## Divider ##########################
    bpy.types.Scene.rename_key_column = bpy.props.IntProperty(
        name="键列",
        default=2,
        min=1,
    )
    bpy.types.Scene.rename_value_column = bpy.props.IntProperty(
        name="值列",
        default=1,
        min=1,
    )
    bpy.types.Scene.rename_save_column = bpy.props.IntProperty(
        name="源保留骨骼列",
        default=3,
        min=1,
    )
    bpy.types.Scene.rename_target_save_column = bpy.props.IntProperty(
        name="目标保留骨骼列",
        default=6,
        min=1,
    )

def unregister():
    bpy.utils.unregister_class(O_ImportSimpleExcel)
    bpy.utils.unregister_class(O_BoneSimpleMapping)
    bpy.utils.unregister_class(O_ImportPosExcel)
    bpy.utils.unregister_class(O_BonePosMapping)
    bpy.utils.unregister_class(O_ImportRenameExcel)
    bpy.utils.unregister_class(O_BoneRenameMapping)
    bpy.utils.unregister_class(P_BoneMapping)

    del bpy.types.Scene.simple_source_armature
    del bpy.types.Scene.source_armature
    del bpy.types.Scene.target_armature
    del bpy.types.Scene.rename_source_mesh
    del bpy.types.Scene.rename_source_armature
    del bpy.types.Scene.rename_target_armature
    

    del bpy.types.Scene.simple_main_column
    del bpy.types.Scene.simple_save_column
    del bpy.types.Scene.simple_toactive_column
    del bpy.types.Scene.simple_active_column

    del bpy.types.Scene.key_column
    del bpy.types.Scene.value_column
    del bpy.types.Scene.rename_key_column
    del bpy.types.Scene.rename_value_column
    del bpy.types.Scene.rename_save_column
    del bpy.types.Scene.rename_target_save_column


